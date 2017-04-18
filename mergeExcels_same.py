""" Created on 18/04/2017

    @author: RMartinezElizalde

    @bried: Merge Excel sheets by row, identifying first element of row (key/index), which its value is a string ending with some numbers
    In this case, sheets are sorted by these numbers, so we extract them and sort the working lists and final merged list
    'same' in script name means result sheet is in the same Excel file containing the original sheets
    """

from openpyxl import load_workbook
import re

# Load in the workbook
wb = load_workbook('..\\..\\test.xlsx')

# Get sheet names
print(wb.get_sheet_names())

# Get a sheet by name
sheet3 = wb.get_sheet_by_name('Sheet3')
sheet4 = wb.get_sheet_by_name('Sheet4')

result = wb.get_sheet_by_name('result')


""" From first columns, get digits in ERRATA string; doing so, we'll be able to sort the list"""
idx = 3
sheet3List = []
sheet4List = []
while True:
    v3 = sheet3.cell(row=idx, column=1).value
    a = ""
    if v3 != None:
        x = []
        for s in v3:
            x.append(re.findall("[-+]?\d+[\.]?\d*[eE]?[-+]?\d*", s))
        for i in x:
            if i != []:
                a = a + i[0]

    v4 = sheet4.cell(row=idx, column=1).value
    b = ""
    if v4 != None:
        y = []
        for s in v4:
            y.append(re.findall("[-+]?\d+[\.]?\d*[eE]?[-+]?\d*", s))
        for i in y:
            if i != []:
                b = b + i[0]

    if a == "" and b == "":
        break

    """ Save digits of ERRATA name with their corresponding row index """
    if a != "":
        sheet3List.append([a, idx])
    if b != "":
        sheet4List.append([b, idx])

    idx = idx + 1

print sheet3List
print sheet4List

""" Get new lists sorted with just the ERRATA numbers. UPDATE: there is not need to sort as they are sorted below """
justNumsList3 = []
for i in sheet3List:
    justNumsList3.append(i[0])

justNumsList4 = []

for i in sheet4List:
    justNumsList4.append(i[0])

print justNumsList3
print justNumsList4

list33 = []
for i in justNumsList3:
    list33.append(int(i))

print list33
list33 = sorted(list33)
print list33

list333 = [] #this is just for removing dupplicated entries (NOT likely to happen)
for i in list33:
    if i not in list333:
        list333.append(i)

print list333

list44 = []
for i in justNumsList4:
    list44.append(int(i))

print list44
list44 = sorted(list44)
print list44

list444 = []
for i in list44:
    if i not in list444:
        list444.append(i)

print list444


""" Merge both errata numbers lists and sort the final list"""
mergedList = []
for i in list333:
    mergedList.append(i)

for i in list444:
    if i not in mergedList: #again, do not dupplicate elements
        mergedList.append(i)

print "len list333: " + str(len(list333))
print "len list444: " + str(len(list444))
print "len mergedList: " + str(len(mergedList))

mergedList = sorted(mergedList)

new_rows = []

""" For each errata number, look for it in the lists (one for each original sheet) created above which contains the row number. So, we can read the original row and write it to the output sheet """
""" If errata number is not found in the first sheet, we look for it in the second sheet. If it is not found in either, we raise an exception! """
idx2 = 2
for i in mergedList:
    match=False
    for j in sheet3List:
        if i == int(j[0]):
            match=True
            for c in sheet3[j[1]]:
                result.cell(column=sheet3[j[1]].index(c)+1, row=idx2, value=c.value)
    if match == False:
        for j in sheet4List:
            if i == int(j[0]):
                match=True
                for c in sheet4[j[1]]:
                    result.cell(column=sheet4[j[1]].index(c)+1, row=idx2, value=c.value)
    if match == False:
        raise Exception("Not found")

    idx2 = idx2 + 1


wb.save('..\\..\\test.xlsx')







